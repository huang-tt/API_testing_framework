import os,sys
dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(dir)	#添加python项目的环境地址
sys.path.append("F:\Python\Lib\site-packages")#添加python项目的第三库的环境地址

#封装读取文件方法
import xlrd
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, colors
from common.read_file import *

class OperationExcel(object):
    #获取sheet表
    def getSheet(self):
        #打开Excel文件
        book = xlrd.open_workbook(filePath())

        #根据索引获取sheet表
        return book.sheet_by_index(0)


    #以列表形式读取出所有数据
    def getExcelDATAS(self):
        data = []
        #获取第一行
        # title = self.getSheet().row_values(0)
        #将第一行改为英文
        title = firstline.values()
        # print(title)
        #从第二行开始循环读取
        for row in range(1,self.getSheet().nrows):
            #获取每一行的内容
            row_value = self.getSheet().row_values(row)
            #将读取出来的每个用例作为一个字典添加到data列表
            data.append(dict(zip(title,row_value)))
        # for i in data:
        #     data = i
        return data

    def write(self,data):
        """
        将数据写入excel表
        :param data: 数据
        :param sheet_name: 自定义excel表名
        :return:
        """
        # 创建一个新的xlsx文件
        workbook = openpyxl.load_workbook("F:\\Python\\Test_learn\\API_testing_framework\\data\\result.xlsx")
        #打开表单
        sheet = workbook['Sheet1']

        #查找用例ID
        rows = int(data['case_id'])

        #将原有Excel的数据，写入新建的Excel中
        va = list(data.items())
        # print("--------va的值",va)
        #默认行和列都从0开始
        row = 1
        col = 1
        #通过循环分别找出键和值
        for item, cost in (va):
            # print("item的值",item)
            # print('cost的值',cost)
            #通过行和列，写入值
            sheet.cell(row, col, item)
            sheet.cell(row+rows, col, cost)
            col += 1
            #通过行和列，获取单元格坐标，例如：“M3”
            coordinate = str('M'+str(int(row)+int(rows)))
            #对比该坐标的值，如果是通过，显示绿色；如果失败，显示红色；没有，则显示黄色
            if sheet[coordinate].value == "Pass":
                sheet[coordinate].fill = PatternFill("solid", fgColor="00EE00")
            elif sheet[coordinate].value == "Fail":
                sheet[coordinate].fill = PatternFill("solid", fgColor="FF0000")
            else:
                sheet[coordinate].fill = PatternFill("solid", fgColor="FFFF00")

        #保存并关闭文件
        workbook.save("F:\\Python\\Test_learn\\API_testing_framework\\data\\result.xlsx")










#对Excel第一行进行更改
firstline = {
    '用例ID': 'case_id',
    '用例模块': 'case_module',
    '用例名称': 'case_name',
    '用例地址': 'case_url',
    '请求方式': 'case_method',
    '请求类型': 'case_type',
    '请求参数': 'case_data',
    '请求头': 'case_headers',
    '前置条件': 'case_preposition',
    '是否执行': 'case_isRun',
    '状态码': 'case_code',
    '断言': 'case_assert',
    '是否通过':'case_result',
    '备注':'case_remarks'
}



excel = OperationExcel().getExcelDATAS()
print(excel)
for i in excel:
    data = i
print(data)





# print(excel)
# print(data)


# print(excel)

